from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path
import json
import math
import re
import statistics
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
ORDER_FILE = ROOT / "近30天出库订单.xlsx"
INVENTORY_FILE = ROOT / "美西库存数据.xlsx"
INBOUND_FILE = ROOT / "近30天入库数据.xlsx"
OUT_DIR = ROOT / "melo"
OUT_JSON = OUT_DIR / "melo_analysis.json"

NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}

SOP = {
    "source_url": "https://ops.wpop.us/projects/melo-sop",
    "sop_date": "2026-06-29",
    "customer": "MLW01",
    "current_warehouse": "WPLA16",
    "future_warehouse_assumption": "集中到一个仓库后，按 LA18 / 新集中仓落地设计",
    "thresholds": {
        "hot_sku_daily_units": 30,
        "beta_daily_orders": 10,
        "gamma_wave_orders": 50,
        "alpha_uph": "120-150 件/人/小时",
    },
}


def col_to_num(ref: str) -> int:
    letters = re.match(r"[A-Z]+", ref).group(0)
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch) - 64
    return n


def load_shared_strings(zf: zipfile.ZipFile):
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    out = []
    with zf.open("xl/sharedStrings.xml") as fh:
        for event, elem in ET.iterparse(fh, events=("end",)):
            if elem.tag.endswith("}si"):
                parts = [t.text or "" for t in elem.findall(".//a:t", NS)]
                out.append("".join(parts))
                elem.clear()
    return out


def cell_value(cell, shared):
    t = cell.attrib.get("t")
    if t == "inlineStr":
        parts = [x.text or "" for x in cell.findall(".//a:t", NS)]
        return "".join(parts)
    v = cell.find("a:v", NS)
    if v is None or v.text is None:
        return ""
    if t == "s":
        return shared[int(v.text)]
    return v.text


def read_header(path: Path, sheet_xml="xl/worksheets/sheet1.xml"):
    with zipfile.ZipFile(path) as zf:
        shared = load_shared_strings(zf)
        with zf.open(sheet_xml) as fh:
            for event, elem in ET.iterparse(fh, events=("end",)):
                if elem.tag.endswith("}row"):
                    values = {}
                    for cell in elem.findall("a:c", NS):
                        values[col_to_num(cell.attrib["r"])] = cell_value(cell, shared)
                    elem.clear()
                    max_col = max(values) if values else 0
                    return [values.get(i, "") for i in range(1, max_col + 1)]
    return []


def iter_target_rows(path: Path, target_names, sheet_xml="xl/worksheets/sheet1.xml"):
    header = read_header(path, sheet_xml)
    wanted = {i + 1: name for i, name in enumerate(header) if name in target_names}
    with zipfile.ZipFile(path) as zf:
        shared = load_shared_strings(zf)
        with zf.open(sheet_xml) as fh:
            for event, elem in ET.iterparse(fh, events=("end",)):
                if not elem.tag.endswith("}row"):
                    continue
                row_num = int(elem.attrib.get("r", "0"))
                if row_num == 1:
                    elem.clear()
                    continue
                row = {}
                for cell in elem.findall("a:c", NS):
                    col = col_to_num(cell.attrib["r"])
                    name = wanted.get(col)
                    if name:
                        row[name] = cell_value(cell, shared)
                elem.clear()
                yield row


def as_float(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", ""))
    except Exception:
        return default


def as_int(v, default=0):
    return int(round(as_float(v, default)))


def parse_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if not v:
        return None
    text = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def day_key(v):
    dt = parse_dt(v)
    return dt.date().isoformat() if dt else "未知"


def hours_between(start, end):
    a = parse_dt(start)
    b = parse_dt(end)
    if not a or not b:
        return None
    return max(0, (b - a).total_seconds() / 3600)


def pct(part, whole, digits=1):
    return round(part / whole * 100, digits) if whole else 0


def percentile(values, q):
    vals = sorted(v for v in values if isinstance(v, (int, float)) and not math.isnan(v))
    if not vals:
        return 0
    pos = (len(vals) - 1) * q / 100
    lo = math.floor(pos)
    hi = math.ceil(pos)
    if lo == hi:
        return round(vals[int(pos)], 2)
    return round(vals[lo] * (hi - pos) + vals[hi] * (pos - lo), 2)


def order_day(order_code: str):
    m = re.search(r"-(\d{6})-", str(order_code))
    if not m:
        return "未知"
    s = m.group(1)
    return f"20{s[:2]}-{s[2:4]}-{s[4:6]}"


def sku_items(row):
    out = []
    for n in range(1, 15):
        sku = row.get(f"SKU{n}-条码/Barcode", "")
        if not sku:
            continue
        out.append(
            {
                "sku": sku,
                "qty": as_int(row.get(f"SKU{n}-数量/Quantity"), 1),
                "weight": as_float(row.get(f"SKU{n}-重量(KG)/Weight")),
                "volume": as_float(row.get(f"SKU{n}-总体积(方)/totalVolume(M³)")),
                "length": as_float(row.get(f"SKU{n}-长度/Length")),
                "width": as_float(row.get(f"SKU{n}-宽度/Width")),
                "height": as_float(row.get(f"SKU{n}-高度/Height")),
            }
        )
    return out


def classify_family(l, w, h, weight):
    longest = max(l, w, h)
    shortest = min(l, w, h)
    if weight >= 4.5 or longest >= 52 or (w >= 28 and h >= 28):
        return "大枕头/大抱枕"
    if shortest <= 5 or h <= 5:
        return "平面件/垫类"
    if weight >= 1.5 or longest >= 35:
        return "中枕头/床品"
    return "小件/轻抛"


def load_inventory():
    wb = load_workbook(INVENTORY_FILE, read_only=True, data_only=True)
    ws = wb.active
    header = [v for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    def g(row, name):
        return row[idx[name]] if name in idx and idx[name] < len(row) else ""

    totals = Counter()
    families = Counter()
    storage = Counter()
    by_sku = {}
    rows = []
    ages = []
    max_sides = []
    locations = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(g(row, "产品代码") or "").strip()
        if not sku:
            continue
        l = as_float(g(row, "长-L"))
        w = as_float(g(row, "宽-W"))
        h = as_float(g(row, "高-H"))
        weight = as_float(g(row, "重-kg"))
        family = classify_family(l, w, h, weight)
        obj = {
            "sku": sku,
            "warehouse": g(row, "仓库"),
            "qty": as_int(g(row, "库存数量-QTY")),
            "pallets": as_float(g(row, "托数-Pallets")),
            "locations": as_int(g(row, "库位数量-Locations")),
            "volume": as_float(g(row, "总体积-VOL")),
            "length": l,
            "width": w,
            "height": h,
            "weight": weight,
            "family": family,
            "storage": g(row, "建议-Target") or g(row, "当前存储类型-Current"),
            "recent_7_out": as_int(g(row, "近7天ToB出库件数")) + as_int(g(row, "近7天ToC出库件数")),
            "recent_60_out": as_int(g(row, "近60天ToB出库件数")) + as_int(g(row, "近60天ToC出库件数")),
            "recent_60_in": as_int(g(row, "近60天入库件数")),
            "max_age": as_int(g(row, "最大库龄-Max-age")),
        }
        by_sku[sku] = obj
        rows.append(obj)
        totals["skus"] += 1
        totals["qty"] += obj["qty"]
        totals["pallets"] += obj["pallets"]
        totals["locations"] += obj["locations"]
        totals["volume"] += obj["volume"]
        totals["recent_7_out"] += obj["recent_7_out"]
        totals["recent_60_out"] += obj["recent_60_out"]
        totals["recent_60_in"] += obj["recent_60_in"]
        families[family] += 1
        storage[str(obj["storage"])] += 1
        ages.append(obj["max_age"])
        max_sides.append(max(l, w, h))
        locations.append(obj["locations"])
    wb.close()
    return {
        "totals": dict(totals),
        "family_counter": dict(families),
        "storage_counter": dict(storage),
        "top_inventory": sorted(rows, key=lambda x: x["qty"], reverse=True)[:15],
        "top_volume": sorted(rows, key=lambda x: x["volume"], reverse=True)[:15],
        "dimension_summary": {
            "max_side_p90_cm": percentile(max_sides, 90),
            "max_age_p90_days": percentile(ages, 90),
            "locations_p90": percentile(locations, 90),
        },
        "by_sku": by_sku,
    }


def load_inbound():
    wb = load_workbook(INBOUND_FILE, read_only=True, data_only=True)
    ws = wb.active
    header = [v for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    def g(row, name):
        return row[idx[name]] if name in idx and idx[name] < len(row) else ""

    totals = Counter()
    warehouses = Counter()
    status_receiving = Counter()
    status_putaway = Counter()
    day_received = Counter()
    day_arrival = Counter()
    sku_qty = Counter()
    sku_lines = Counter()
    receiving_codes = set()
    containers = set()
    arrival_to_unload = []
    unload_to_receive = []
    receive_to_putaway = []
    arrival_to_putaway = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(g(row, "产品代码/Product code") or "").strip()
        receiving_code = str(g(row, "入库单号/Receiving code") or "").strip()
        tracking = str(g(row, "跟踪号/Tracking number") or "").strip()
        expected = as_int(g(row, "预期总件数/Expected quantity"))
        received = as_int(g(row, "已收总件数/Quantity received"))
        putaway = as_int(g(row, "上架总件数/Putaway number"))
        volume = as_float(g(row, "预报总体积/Forecast total volume"))
        actual_arrival = g(row, "实际到仓时间(仓库当地时间)/Actual Arrival Time(Local)") or g(row, "实际到仓时间/Actual Arrival Time")
        unload_time = g(row, "卸货时间(仓库当地时间)/Warehouse Local Time") or g(row, "卸货时间/Unloaded Time")
        receive_finish = g(row, "收货完成时间(仓库本地时间)/Finish Receiving Time(Local)") or g(row, "收货完成时间/Receiving Finished Time")
        putaway_finish = g(row, "上架完成时间(仓库本地时间)/Finish Putaway Time(Local)") or g(row, "上架完成时间/finish Putaway Time")

        totals["lines"] += 1
        totals["expected_qty"] += expected
        totals["received_qty"] += received
        totals["putaway_qty"] += putaway
        totals["forecast_volume"] += volume
        totals["mixed_qty"] += as_int(g(row, "混箱数量/Mixed Qty"))
        totals["unpacking_qty"] += as_int(g(row, "拆箱数量/Unpacking Qty"))
        if receiving_code:
            receiving_codes.add(receiving_code)
        if tracking:
            containers.add(tracking)
        warehouses[str(g(row, "仓库代码/Warehouse") or "未知")] += 1
        status_receiving[str(g(row, "收货状态/Receiving status") or "未知")] += 1
        status_putaway[str(g(row, "上架状态/Putaway status") or "未知")] += 1
        day_received[day_key(receive_finish)] += received
        day_arrival[day_key(actual_arrival)] += received
        if sku:
            sku_qty[sku] += received
            sku_lines[sku] += 1

        for bucket, start, end in [
            (arrival_to_unload, actual_arrival, unload_time),
            (unload_to_receive, unload_time, receive_finish),
            (receive_to_putaway, receive_finish, putaway_finish),
            (arrival_to_putaway, actual_arrival, putaway_finish),
        ]:
            h = hours_between(start, end)
            if h is not None:
                bucket.append(h)

    wb.close()
    totals["receiving_orders"] = len(receiving_codes)
    totals["containers"] = len(containers)
    active_days = [d for d in day_received if d != "未知"]
    top_skus = [{"sku": sku, "received_qty": qty, "lines": sku_lines[sku]} for sku, qty in sku_qty.most_common(15)]

    return {
        "totals": dict(totals),
        "warehouses": warehouses.most_common(),
        "status_receiving": status_receiving.most_common(),
        "status_putaway": status_putaway.most_common(),
        "daily_received": [{"date": d, "qty": day_received[d]} for d in sorted(active_days)],
        "daily_arrival": [{"date": d, "qty": day_arrival[d]} for d in sorted(d for d in day_arrival if d != "未知")],
        "top_skus": top_skus,
        "stats": {
            "arrival_to_unload_p50_h": percentile(arrival_to_unload, 50),
            "unload_to_receive_p50_h": percentile(unload_to_receive, 50),
            "receive_to_putaway_p50_h": percentile(receive_to_putaway, 50),
            "arrival_to_putaway_p50_h": percentile(arrival_to_putaway, 50),
            "arrival_to_putaway_p90_h": percentile(arrival_to_putaway, 90),
            "avg_daily_received_qty": round(sum(day_received[d] for d in active_days) / max(1, len(active_days))),
        },
    }


def target_columns():
    cols = [
        "订单号/Order Code",
        "客户代码/CustomerCode",
        "仓库/Warehouse",
        "订单类型/Order Type",
        "销售平台/Platform",
        "拣货类型/Picking Type",
        "件数/Quantity",
        "物流渠道/Logistics Channel",
        "州/省/Recipients Province/State",
        "推荐包材/Recommended Packaging Materials",
        "出库时效(小时-工作日)/Outbound Time prescription(hour)",
    ]
    for n in range(1, 15):
        cols += [
            f"SKU{n}-条码/Barcode",
            f"SKU{n}-数量/Quantity",
            f"SKU{n}-重量(KG)/Weight",
            f"SKU{n}-总体积(方)/totalVolume(M³)",
            f"SKU{n}-长度/Length",
            f"SKU{n}-宽度/Width",
            f"SKU{n}-高度/Height",
        ]
    return set(cols)


def first_pass():
    daily_single_multi = Counter()
    for row in iter_target_rows(ORDER_FILE, target_columns()):
        items = sku_items(row)
        if len(items) == 1 and items[0]["qty"] >= 2:
            daily_single_multi[(order_day(row.get("订单号/Order Code", "")), items[0]["sku"])] += 1
    return daily_single_multi


def route_for(items, day, daily_single_multi):
    if len(items) == 1 and items[0]["qty"] == 1:
        return "Alpha 单品单件快道"
    if len(items) == 1 and items[0]["qty"] >= 2:
        if daily_single_multi[(day, items[0]["sku"])] > SOP["thresholds"]["beta_daily_orders"]:
            return "Beta 单品多件工作站"
        return "Gamma 播种墙"
    return "Gamma 播种墙"


def analyze_orders(inv_by_sku, daily_single_multi):
    totals = Counter()
    route_orders = Counter()
    route_units = Counter()
    route_days = defaultdict(Counter)
    picking = Counter()
    channel = Counter()
    platform = Counter()
    carrier = Counter()
    state = Counter()
    packaging = Counter()
    warehouse = Counter()
    day_orders = Counter()
    day_units = Counter()
    day_structure = defaultdict(Counter)
    sku_units = Counter()
    sku_orders = Counter()
    combos = Counter()
    order_units = []
    sku_counts = []
    outbound_hours = []

    for row in iter_target_rows(ORDER_FILE, target_columns()):
        items = sku_items(row)
        units = sum(i["qty"] for i in items) or as_int(row.get("件数/Quantity"))
        day = order_day(row.get("订单号/Order Code", ""))
        route = route_for(items, day, daily_single_multi)
        if len(items) == 1 and items[0]["qty"] == 1:
            structure = "单品单件"
        elif len(items) == 1 and items[0]["qty"] >= 2:
            structure = "单品多件"
        else:
            structure = "多品多件"
        totals["orders"] += 1
        totals["units"] += units
        totals["sku_lines"] += len(items)
        route_orders[route] += 1
        route_units[route] += units
        route_days[route][day] += 1
        picking[row.get("拣货类型/Picking Type", "未知") or "未知"] += 1
        channel[row.get("订单类型/Order Type", "未知") or "未知"] += 1
        platform[row.get("销售平台/Platform", "未知") or "未知"] += 1
        carrier[row.get("物流渠道/Logistics Channel", "未知") or "未知"] += 1
        state[row.get("州/省/Recipients Province/State", "未知") or "未知"] += 1
        packaging[row.get("推荐包材/Recommended Packaging Materials", "未推荐") or "未推荐"] += 1
        warehouse[row.get("仓库/Warehouse", "未知") or "未知"] += 1
        day_orders[day] += 1
        day_units[day] += units
        day_structure[day][structure] += 1
        order_units.append(units)
        sku_counts.append(len(items))
        oh = as_float(row.get("出库时效(小时-工作日)/Outbound Time prescription(hour)"), None)
        if oh is not None and oh >= 0:
            outbound_hours.append(oh)
        if len(items) >= 2:
            combos[" + ".join(sorted(i["sku"] for i in items)[:4])] += 1
        for item in items:
            sku_units[item["sku"]] += item["qty"]
            sku_orders[item["sku"]] += 1

    active_days = [d for d in day_orders if d != "未知"]
    days = max(1, len(active_days))
    top_skus = []
    for sku, units in sku_units.most_common(25):
        inv = inv_by_sku.get(sku, {})
        avg = units / days
        stock = inv.get("qty", 0)
        top_skus.append(
            {
                "sku": sku,
                "orders": sku_orders[sku],
                "units_30d": units,
                "avg_daily_units": round(avg, 1),
                "stock_qty": stock,
                "stock_days": round(stock / avg, 1) if avg else 0,
                "family": inv.get("family", "未匹配"),
                "dims": f"{inv.get('length', 0):g}×{inv.get('width', 0):g}×{inv.get('height', 0):g} cm",
                "weight": inv.get("weight", 0),
            }
        )

    return {
        "totals": dict(totals),
        "route_orders": dict(route_orders),
        "route_units": dict(route_units),
        "route_daily_avg_orders": {r: round(sum(c.values()) / max(1, len(c)), 1) for r, c in route_days.items()},
        "picking_type": picking.most_common(),
        "channel_orders": channel.most_common(),
        "top_platforms": platform.most_common(8),
        "top_carriers": carrier.most_common(12),
        "top_states": state.most_common(15),
        "packaging": packaging.most_common(12),
        "warehouses": warehouse.most_common(),
        "daily": [{"date": d, "orders": day_orders[d], "units": day_units[d]} for d in sorted(active_days)],
        "daily_structure": [
            {
                "date": d,
                "single_one": day_structure[d]["单品单件"],
                "single_multi": day_structure[d]["单品多件"],
                "multi_sku": day_structure[d]["多品多件"],
                "total": sum(day_structure[d].values()),
            }
            for d in sorted(active_days)
        ],
        "top_skus": top_skus,
        "top_combinations": combos.most_common(12),
        "stats": {
            "order_units_p50": percentile(order_units, 50),
            "order_units_p90": percentile(order_units, 90),
            "sku_count_p50": percentile(sku_counts, 50),
            "sku_count_p90": percentile(sku_counts, 90),
            "outbound_hours_median": round(statistics.median(outbound_hours), 2) if outbound_hours else 0,
            "outbound_hours_p90": percentile(outbound_hours, 90),
        },
    }


def build_design(data):
    orders = data["orders"]
    inv = data["inventory"]
    inbound = data["inbound"]
    total_orders = orders["totals"]["orders"]
    routes = orders["route_orders"]
    route_line = " / ".join(f"{k.replace(' 单品单件快道','').replace(' 单品多件工作站','').replace(' 播种墙','')} {pct(v, total_orders)}%" for k, v in routes.items())
    return {
        "diagnosis": [
            f"近30天出库 {total_orders:,} 单、{orders['totals']['units']:,} 件，日均约 {round(total_orders / 30):,} 单；集中仓需要按高峰日 8,000+ 单预留弹性。",
            f"按页面 SOP 阈值重算三通道：{route_line}。单品订单占主导，播种墙不应承接全部订单。",
            f"库存 {inv['totals']['skus']} 个 SKU、{inv['totals']['qty']:,} 件、约 {round(inv['totals']['pallets']):,} 托、{round(inv['totals']['volume']):,} m³；库存深度高，适合做前置区+后备区双层库存。",
            f"近30天入库 {inbound['totals']['received_qty']:,} 件、{inbound['totals']['containers']} 个柜/跟踪号，上架完成中位时效 {inbound['stats']['arrival_to_putaway_p50_h']} 小时；集中仓需要独立收货缓冲和上架优先级。",
            "SKU 体积以枕头/床品为主，库内瓶颈更像“大件轻抛搬运 + 播种格口占用”，而不是传统小件密集拣选。",
        ],
        "zones": [
            {"name": "R 收货缓冲与质检区", "size": "靠入库 dock，按柜号/入库单划临时 staging", "logic": "卸柜、点数、异常差异、待上架托盘分流；爆款 SKU 到货后优先补 A 区。"},
            {"name": "A 爆款前置整托区", "size": "靠近 Dock，Top10-25 SKU，每 SKU 0.5-6 托", "logic": "承接 Alpha 和高频 Beta，整托下架、扫码到托、贴标原包装出库。"},
            {"name": "B 单品多件工作站", "size": "2-4 个双人工位，配分格货架", "logic": "同 SKU 多件且日订单 >10 单独立处理，1人分货、1人打包贴标。"},
            {"name": "C 多品播种区", "size": "4-6 面开放式可调层板播种墙，每墙约50格", "logic": "Gamma 50单/波次，按承运商 cutoff、SKU 重叠度、大件订单分波。"},
            {"name": "D 大件/异常区", "size": "地面画格 + 大格位货架 + 缺货/破损暂存", "logic": "55×30×30cm 级大件独立波次；超规格包材订单组长处理。"},
            {"name": "E 包材与压缩区", "size": "5种标准包材 7天安全库存", "logic": "小袋/大袋/小中大箱，系统推荐只允许升档；预留真空压缩位降低 DIM weight。"},
            {"name": "F 出库分拨暂存", "size": "按 USPS / Amazon Ground / FedEx / Other 分 lane", "logic": "三通道汇合后复核、打托、装车，按 pickup cutoff 管控优先级。"},
        ],
        "process": [
            {"step": "0. 入库收货", "text": "按柜号/入库单卸货，点数收货后按 SKU 热度决定补前置区、进后备区或进异常区。"},
            {"step": "1. 订单池分流", "text": "WMS 按 SKU 种类、件数、当日 SKU 单量打 Alpha/Beta/Gamma 标签。"},
            {"step": "2. Alpha 快道", "text": "一票一件直接从爆款前置区或普通位批量拣，扫托盘/扫件后贴标出库。"},
            {"step": "3. Beta 工作站", "text": "单 SKU 多件批量下架到站，面单标组包数，分格后捆扎/装箱。"},
            {"step": "4. Gamma 播种", "text": "多品订单生成 50 单波次，汇总拣货、扫码亮灯/显示格口、齐套后打包。"},
            {"step": "5. 异常闭环", "text": "缺货、破损、包材不匹配、超规格分别进异常区，系统挂起并日报。"},
            {"step": "6. 出库交接", "text": "按承运商 lane 暂存，截单前集中复核打托装车。"},
        ],
        "staffing": [
            {"area": "入库/上架", "staff": "3-6人", "target": "到仓后当日收货，爆款优先上架"},
            {"area": "Alpha", "staff": "2-4人", "target": "120-150 件/人/小时"},
            {"area": "Beta", "staff": "2-4人", "target": "每站2人，优先高频 SKU"},
            {"area": "Gamma", "staff": "8-12人", "target": "4-6面墙并行，每墙2人"},
            {"area": "补货/异常/打托", "staff": "3-5人", "target": "爆款不断货，异常不过夜"},
        ],
        "open_questions": [
            "集中仓是否最终确定为 LA18？需要 dock 位置、可用面积、柱网和通道宽度后才能画精确比例库图。",
            "WMS 是否已支持“扫码显示目标格口号”、托盘容器精确到托、面单正反面组包数打印？",
            "入库收货是否按柜号、入库单号还是 SKU 维度生成上架任务？这会影响 staging 区划线和优先补货逻辑。",
            "承运商 pickup 截止时间、每天车次和是否混装，会影响出库暂存 lane 数量与波次释放节奏。",
        ],
    }


def main():
    OUT_DIR.mkdir(exist_ok=True)
    inventory = load_inventory()
    inbound = load_inbound()
    daily_single_multi = first_pass()
    orders = analyze_orders(inventory["by_sku"], daily_single_multi)
    data = {
        "sop": SOP,
        "inventory": {k: v for k, v in inventory.items() if k != "by_sku"},
        "inbound": inbound,
        "orders": orders,
    }
    data["design"] = build_design(data)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUT_JSON),
        "orders": orders["totals"],
        "routes": orders["route_orders"],
        "inbound": inbound["totals"],
        "inventory": data["inventory"]["totals"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
