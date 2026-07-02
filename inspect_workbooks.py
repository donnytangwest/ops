from pathlib import Path
import json
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
FILES = [
    "产品基础信息.xlsx",
    "美西库存数据.xlsx",
    "近30天出库订单.xlsx",
]


def summarize_file(path: Path):
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = ws.iter_rows(max_row=6, values_only=True)
            rows = [[("" if v is None else str(v)) for v in row] for row in rows]
            columns = rows[0] if rows else []
            sample = [dict(zip(columns, row)) for row in rows[1:]]
            sheets.append(
                {
                    "sheet": ws.title,
                    "columns": columns,
                    "sample": sample,
                    "row_count": ws.max_row - 1 if ws.max_row else None,
                    "max_column": ws.max_column,
                }
            )
        wb.close()
        return {"file": path.name, "size_bytes": path.stat().st_size, "sheets": sheets}

    xl = pd.ExcelFile(path)
    sheets = []
    for sheet in xl.sheet_names:
        header_df = pd.read_excel(path, sheet_name=sheet, nrows=0)
        sample_df = pd.read_excel(path, sheet_name=sheet, nrows=5)
        sheets.append(
            {
                "sheet": sheet,
                "columns": [str(c) for c in header_df.columns],
                "sample": sample_df.fillna("").astype(str).to_dict(orient="records"),
                "row_count": None,
            }
        )
    return {"file": path.name, "size_bytes": path.stat().st_size, "sheets": sheets}


def main():
    out = []
    for name in FILES:
        path = ROOT / name
        out.append(summarize_file(path))
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
