from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path
import json
import math
import statistics

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
ORDER_FILE = ROOT / "近30天出库订单.xlsx"
INVENTORY_FILE = ROOT / "美西库存数据.xlsx"
PRODUCT_FILE = ROOT / "产品基础信息.xlsx"
OUT_DIR = ROOT / "site"
OUT_JSON = OUT_DIR / "melo_analysis.json"

SOP = {
    "source_url": "https://ops.wpop.us/projects/melo-sop",
    "sop_date": "2026-06-29",
    "warehouse_now": "LA16",
    "future_warehouse_assumption": "集中至单一仓库，页面方案按 LA18 / 新集中仓实施假设设计",
    "thresholds": {
        "hot_sku_daily_units": 30,
        "beta_single_sku_multi_order_daily_orders": 10,
        "gamma_wave_orders": 50,
        "alpha_uph": "120-150 件/人/小时",
        "beta_staff_per_station": 2,
        "gamma_staff_per_station": 2,
    },
}


def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def as_float(v, default=0.0):
    v = norm(v)
    if v == "":
        return default
    try:
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except Exception:
        return default


def as_int(v, default=0):
    return int(round(as_float(v, default)))


def parse_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str) and v.strip():
        text = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text[:19], fmt).date().isoformat()
            except Exception:
                pass
        return text[:10]
    return "未知"


def pct(part, whole, digits=1):
    if not whole:
        return 0
    return round(part / whole * 100, digits)


def read_header(ws):
    return [norm(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def index_header(header):
    return {name: i for i, name in enumerate(header)}


def get(row, idx, name, default=""):
    pos = idx.get(name)
    if pos is None or pos >= len(row):
        return default
    return norm(row[pos])


def order_skus(row, idx):
    items = []
    for n in range(1, 15):
        sku = get(row, idx, f"SKU{n}-条码/Barcode")
        if not sku:
            continue
        qty = as_int(get(row, idx, f"SKU{n}-数量/Quantity"), 1)
        weight = as_float(get(row, idx, f"SKU{n}-重量(KG)/Weight"))
        volume = as_float(get(row, idx, f"SKU{n}-总体积(方)/totalVolume(M³)"))
        length = as_float(get(row, idx, f"SKU{n}-长度/Length"))
        width = as_float(get(row, idx, f"SKU{n}-宽度/Width"))
        height = as_float(get(row, idx, f"SKU{n}-高度/Height"))
        items.append(
            {
                "sku": str(sku),
                "qty": qty,
                "weight": weight,
                "volume": volume,
                "length": length,
                "width": width,
                "height": height,
            }
        )
    return items


def load_inventory():
    wb = load_workbook(INVENTORY_FILE, read_only=True, data_only=True)
    ws = wb.active
    header = read_header(ws)
    idx = index_header(header)
    rows = []
    by_sku = {}
    totals = Counter()
    dims = []
    age_values = []
    location_counts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(get(row, idx, "产品代码"))
        if not sku:
            continue
        qty = as_int(get(row, idx, "库存数量-QTY"))
        pallets = as_float(get(row, idx, "托数-Pallets"))
        volume = as_float(get(row, idx, "总体积-VOL"))
        locs = as_int(get(row, idx, "库位数量-Locations"))
        recent_toc_7 = as_int(get(row, idx, "近7天ToC出库件数"))
        recent_tob_7 = as_int(get(row, idx, "近7天ToB出库件数"))
        recent_in_7 = as_int(get(row, idx, "近7天入库件数"))
        recent_toc_60 = as_int(get(row, idx, "近60天ToC出库件数"))
        recent_tob_60 = as_int(get(row, idx, "近60天ToB出库件数"))
        recent_in_60 = as_int(get(row, idx, "近60天入库件数"))
        l = as_float(get(row, idx, "长-L"))
        w = as_float(get(row, idx, "宽-W"))
        h = as_float(get(row, idx, "高-H"))
        weight = as_float(get(row, idx, "重-kg"))
        max_age = as_int(get(row, idx, "最大库龄-Max-age"))
        family = classify_family(l, w, h, weight)
        row_obj = {
            "sku": sku,
            "warehouse": get(row, idx, "仓库"),
            "qty": qty,
            "pallets": pallets,
            "volume": volume,
            "locations": locs,
            "length": l,
            "width": w,
            "height": h,
            "weight": weight,
            "family": family,
            "current_storage": get(row, idx, "当前存储类型-Current"),
            "target_storage": get(row, idx, "建议-Target"),
            "recent_7_out": recent_toc_7 + recent_tob_7,
            "recent_7_in": recent_in_7,
            "recent_60_out": recent_toc_60 + recent_tob_60,
            "recent_60_in": recent_in_60,
            "max_age": max_age,
        }
        rows.append(row_obj)
        by_sku[sku] = row_obj
        totals["skus"] += 1
        totals["qty"] += qty
        totals["pallets"] += pallets
        totals["volume"] += volume
        totals["locations"] += locs
        totals["recent_7_out"] += row_obj["recent_7_out"]
        totals["recent_7_in"] += recent_in_7
        totals["recent_60_out"] += row_obj["recent_60_out"]
        totals["recent_60_in"] += recent_in_60
        dims.append(max(l, w, h))
        age_values.append(max_age)
        location_counts.append(locs)
    wb.close()

    family_counter = Counter(r["family"] for r in rows)
    storage_counter = Counter(r["target_storage"] or r["current_storage"] for r in rows)
    top_inventory = sorted(rows, key=lambda r: r["qty"], reverse=True)[:12]
    top_volume = sorted(rows, key=lambda r: r["volume"], reverse=True)[:12]
    return {
        "totals": dict(totals),
        "family_counter": dict(family_counter),
        "storage_counter": dict(storage_counter),
        "top_inventory": top_inventory,
        "top_volume": top_volume,
        "dimension_summary": {
            "max_side_p50": percentile(dims, 50),
            "max_side_p90": percentile(dims, 90),
            "max_age_p90": percentile(age_values, 90),
            "locations_p90": percentile(location_counts, 90),
        },
        "by_sku": by_sku,
    }


def classify_family(l, w, h, weight):
    longest = max(l, w, h)
    shortest = min(l, w, h)
    if weight >= 4.5 or longest >= 52 or (w >= 28 and h >= 28):
        return "大枕头/大抱枕"
    if shortest <= 5 or h <= 5:
        return "平面件/垫类"
    if weight >= 1.5 or longest >= 35:
        return "中枕头/床品"
    return "小件/轻抛"


def percentile(values, q):
    vals = sorted(v for v in values if isinstance(v, (int, float)) and not math.isnan(v))
    if not vals:
        return 0
    pos = (len(vals) - 1) * q / 100
    lo = math.floor(pos)
    hi = math.ceil(pos)
    if lo == hi:
        return round(vals[int(pos)], 2)
    return round(vals[lo] * (hi - pos) + vals[hi] * (pos - lo), 2)


def first_pass_single_multi():
    wb = load_workbook(ORDER_FILE, read_only=True, data_only=True)
    ws = wb["订单明细"]
    header = read_header(ws)
    idx = index_header(header)
    daily_single_multi = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        items = order_skus(row, idx)
        if len(items) == 1 and items[0]["qty"] >= 2:
            day = parse_date(get(row, idx, "本地交运出库时间/Local outbound date") or get(row, idx, "创建时间/Add Time"))
            daily_single_multi[(day, items[0]["sku"])] += 1
    wb.close()
    return daily_single_multi


def analyze_orders(inventory_by_sku, daily_single_multi):
    wb = load_workbook(ORDER_FILE, read_only=True, data_only=True)
    ws = wb["订单明细"]
    header = read_header(ws)
    idx = index_header(header)

    totals = Counter()
    picking_type = Counter()
    channel_orders = Counter()
    route_orders = Counter()
    route_units = Counter()
    route_days = defaultdict(Counter)
    day_orders = Counter()
    day_units = Counter()
    sku_units = Counter()
    sku_orders = Counter()
    sku_route_units = defaultdict(Counter)
    state_orders = Counter()
    platform_orders = Counter()
    packaging = Counter()
    carrier_orders = Counter()
    order_qty_values = []
    sku_count_values = []
    outbound_hours = []
    combinations = Counter()
    warehouses = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        totals["orders"] += 1
        items = order_skus(row, idx)
        sku_count = len(items) or as_int(get(row, idx, "sku种类"), 0)
        units = sum(item["qty"] for item in items) or as_int(get(row, idx, "件数/Quantity"), 0)
        day = parse_date(get(row, idx, "本地交运出库时间/Local outbound date") or get(row, idx, "创建时间/Add Time"))
        picking = str(get(row, idx, "拣货类型/Picking Type") or "未知")
        route = classify_route(items, day, daily_single_multi)

        totals["units"] += units
        totals["sku_lines"] += sku_count
        picking_type[picking] += 1
        route_orders[route] += 1
        route_units[route] += units
        route_days[route][day] += 1
        day_orders[day] += 1
        day_units[day] += units
        channel_orders[str(get(row, idx, "订单类型/Order Type") or "未知")] += 1
        state_orders[str(get(row, idx, "州/省/Recipients Province/State") or "未知")] += 1
        platform_orders[str(get(row, idx, "销售平台/Platform") or "未知")] += 1
        carrier_orders[str(get(row, idx, "物流渠道/Logistics Channel") or "未知")] += 1
        packaging[str(get(row, idx, "推荐包材/Recommended Packaging Materials") or "未推荐")] += 1
        warehouses[str(get(row, idx, "仓库/Warehouse") or "未知")] += 1
        order_qty_values.append(units)
        sku_count_values.append(sku_count)
        oh = as_float(get(row, idx, "出库时效(小时-工作日)/Outbound Time prescription(hour)"), None)
        if oh is not None and oh >= 0:
            outbound_hours.append(oh)
        if len(items) >= 2:
            key = "+".join(sorted([i["sku"] for i in items])[:4])
            combinations[key] += 1
        for item in items:
            sku = item["sku"]
            sku_units[sku] += item["qty"]
            sku_orders[sku] += 1
            sku_route_units[route][sku] += item["qty"]

    wb.close()

    top_skus = []
    for sku, units in sku_units.most_common(20):
        inv = inventory_by_sku.get(sku, {})
        avg_daily = units / max(1, len([d for d in day_orders if d != "未知"]))
        stock = inv.get("qty", 0)
        top_skus.append(
            {
                "sku": sku,
                "orders": sku_orders[sku],
                "units_30d": units,
                "avg_daily_units": round(avg_daily, 1),
                "stock_qty": stock,
                "stock_days": round(stock / avg_daily, 1) if avg_daily else 0,
                "family": inv.get("family", "未匹配"),
                "dims": f"{inv.get('length', 0):g}×{inv.get('width', 0):g}×{inv.get('height', 0):g} cm",
                "weight": inv.get("weight", 0),
            }
        )

    daily = [
        {"date": d, "orders": day_orders[d], "units": day_units[d]}
        for d in sorted(day_orders)
        if d != "未知"
    ]
    return {
        "totals": dict(totals),
        "picking_type": picking_type.most_common(),
        "route_orders": dict(route_orders),
        "route_units": dict(route_units),
        "route_daily_avg_orders": {
            route: round(sum(days.values()) / max(1, len(days)), 1)
            for route, days in route_days.items()
        },
        "channel_orders": channel_orders.most_common(),
        "top_states": state_orders.most_common(12),
        "top_platforms": platform_orders.most_common(8),
        "top_carriers": carrier_orders.most_common(10),
        "packaging": packaging.most_common(12),
        "warehouses": warehouses.most_common(),
        "top_skus": top_skus,
        "top_combinations": combinations.most_common(12),
        "daily": daily,
        "stats": {
            "order_units_p50": percentile(order_qty_values, 50),
            "order_units_p90": percentile(order_qty_values, 90),
            "sku_count_p50": percentile(sku_count_values, 50),
            "sku_count_p90": percentile(sku_count_values, 90),
            "outbound_hours_median": round(statistics.median(outbound_hours), 2) if outbound_hours else 0,
            "outbound_hours_p90": percentile(outbound_hours, 90),
        },
    }


def classify_route(items, day, daily_single_multi):
    if len(items) == 1 and items[0]["qty"] == 1:
        return "Alpha 单品单件快道"
    if len(items) == 1 and items[0]["qty"] >= 2:
        if daily_single_multi[(day, items[0]["sku"])] > SOP["thresholds"]["beta_single_sku_multi_order_daily_orders"]:
            return "Beta 单品多件工作站"
        return "Gamma 播种墙"
    return "Gamma 播种墙"


def build_design(data):
    route_orders = data["orders"]["route_orders"]
    total_orders = data["orders"]["totals"]["orders"]
    total_units = data["orders"]["totals"]["units"]
    inv = data["inventory"]["totals"]
    alpha_pct = pct(route_orders.get("Alpha 单品单件快道", 0), total_orders)
    beta_pct = pct(route_orders.get("Beta 单品多件工作站", 0), total_orders)
    gamma_pct = pct(route_orders.get("Gamma 播种墙", 0), total_orders)
    return {
        "diagnosis": [
            f"30天出库 {total_orders:,} 单 / {total_units:,} 件，日均约 {round(total_orders / 30):,} 单，已经是高频 DTC 履约体量。",
            f"按 SOP 阈值重分类后，Alpha 占 {alpha_pct}%，Beta 占 {beta_pct}%，Gamma 占 {gamma_pct}%；必须把单品件从播种墙前截流。",
            f"库存侧 {inv.get('skus', 0)} 个 SKU、{inv.get('qty', 0):,} 件、约 {round(inv.get('pallets', 0)):,} 托；SKU 少但库存深，适合集中仓做前置爆款区 + 后备地堆。",
            "货品以枕头/床品为主，轻抛、大体积、箱规敏感；库区设计重点不是重货承载，而是缩短爆款搬运距离和减少大件播种格口冲突。",
        ],
        "zones": [
            {
                "name": "A区 爆款前置整托区",
                "purpose": "承接 Top SKU 与日出 >30 件 SKU，紧贴 dock / 出库暂存。",
                "storage": "地堆整托 + Min/Max 补货位，Top10 SKU 每 SKU 2-6 托前置。",
                "process": "Alpha 单品单件批量扫托盘、打单、贴标、原包装出库；Beta 高频单品多件在旁边货架分格。",
            },
            {
                "name": "B区 单品多件工作站",
                "purpose": "处理同 SKU 多件且当日 >10 单的订单，避免低效占用播种墙。",
                "storage": "SKU 整托到站，工作站配分格货架、打包台、胶带/箱袋。",
                "process": "1人分货到格 + 1人打包贴标；USPS/普通渠道胶带捆扎，FedEx/Amazon 按渠道装箱/袋。",
            },
            {
                "name": "C区 Gamma 播种区",
                "purpose": "处理多品多件和低频单品多件。",
                "storage": "开放式可调层板播种墙，建议 50 格/墙，格口不小于 65×40×40 cm。",
                "process": "50单/波次，按承运商 cutoff、SKU 重叠度、大件订单分层；强制扫码显示格口。",
            },
            {
                "name": "D区 大件/异常处理区",
                "purpose": "隔离 55×30×30 cm 级大件、多品缺货、超规格订单。",
                "storage": "地面画格 + 大格位货架 + 异常暂存。",
                "process": "大件独立波次；超出 5 种标准包材的订单进入组长人工打包。",
            },
            {
                "name": "E区 包材与压缩区",
                "purpose": "降低包材等待与 DIM weight。",
                "storage": "5种标准包材 7天安全库存；可预留真空压缩机位。",
                "process": "系统推荐包材，只允许人工升档；记忆棉可压缩降档。",
            },
            {
                "name": "F区 出库分拨暂存",
                "purpose": "按承运商/渠道/截单时间分堆，减少末端混票。",
                "storage": "USPS、Amazon Ground、FedEx/其他渠道分 lane。",
                "process": "三通道汇合后按渠道打托、复核、装车。",
            },
        ],
        "workforce": [
            {"area": "Alpha 爆款快道", "staff": "2-4人", "target": "120-150 件/人/小时"},
            {"area": "Beta 单品多件", "staff": "2-4人", "target": "按SKU整托分货，优先消化 >10单 SKU"},
            {"area": "Gamma 播种", "staff": "8-12人", "target": "4-6面墙并行，每墙 2人"},
            {"area": "补货/异常/打托", "staff": "3-5人", "target": "补货不断点，异常不过夜"},
        ],
        "open_questions": [
            "集中仓最终是否确定为 LA18？dock 数、可用面积、柱网、消防通道宽度需要确认后才能画精确库位。",
            "WMS 是否已经支持 Gamma 的“扫码显示目标格口号”和 Beta 的正反面组包数打印？",
            "入库 `.xls` 目前因本地缺少老式 Excel 解析库未纳入精确计算，需要你提供转换后的 `.xlsx` 或允许我用可用工具转换后补算。",
            "承运商每日 pickup 截止时间和车次，需要用于确定波次释放节奏与出库暂存 lane 数。",
        ],
    }


def main():
    OUT_DIR.mkdir(exist_ok=True)
    inventory = load_inventory()
    daily_single_multi = first_pass_single_multi()
    orders = analyze_orders(inventory["by_sku"], daily_single_multi)
    data = {
        "sop": SOP,
        "inventory": {k: v for k, v in inventory.items() if k != "by_sku"},
        "orders": orders,
        "design": {},
    }
    data["design"] = build_design(data)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUT_JSON),
        "orders": orders["totals"],
        "routes": orders["route_orders"],
        "inventory": data["inventory"]["totals"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
